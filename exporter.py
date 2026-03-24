"""
Módulo de Exportación - Utilidad
Exporta datos a Excel y PDF
"""

import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors

class DataExporter:
    """Clase para exportar datos a diferentes formatos"""
    
    @staticmethod
    def to_excel(data, filename, headers):
        """
        Exporta datos a Excel
        Parámetros: data (lista de tuplas), filename, headers (lista)
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Hábitos"
        
        # Estilos de cabecera
        header_fill = openpyxl.styles.PatternFill(start_color="4CAF50", 
                                                    end_color="4CAF50", 
                                                    fill_type="solid")
        header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        
        # Escribir cabeceras
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Escribir datos
        for row_num, row_data in enumerate(data, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value if value else "")
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 2
        
        wb.save(filename)
    
    @staticmethod
    def to_pdf(data, filename, title):
        """
        Exporta datos a PDF con formato profesional
        Parámetros: data, filename, title
        """
        doc = SimpleDocTemplate(filename, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = styles['Heading1']
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph("<br/>", styles['Normal']))
        
        # Tabla de datos
        if data:
            # Agregar headers
            table_data = [['Hábito', 'Categoría', 'Completados']]
            for row in data:
                table_data.append([str(row[0]), str(row[1]), str(row[2])])
            
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.green),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
        
        doc.build(elements)